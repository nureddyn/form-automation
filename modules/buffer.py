import PyPDF2
import re
from typing import List, Any, Dict
import csv
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def pdf_buffer(path: str) -> PyPDF2.PdfReader:
    try:
        with open(path, "rb") as f:
            pdf_reader = PyPDF2.PdfReader(f)
            metadata = pdf_reader.metadata

            first_page = pdf_reader.pages[0].extract_text().splitlines()

            # TODO: Remove unnecesary use of regular expression for 'TYPE' field in buffer
            form_type = [line for line in first_page if re.search(r'\b\w*FACILIT\w*\b', line, re.IGNORECASE)][0]
            
            form_fields = {}
            if pdf_reader.get_form_text_fields():
                form_fields =pdf_reader.get_form_text_fields()
            buffer = {'METADATA': metadata, 'reader': pdf_reader, 'TYPE': form_type.strip(), 'fields': form_fields}

    except FileNotFoundError:
        raise FileNotFoundError(f"Error: The file at {path} was not found.")
    except PermissionError:
        raise PermissionError(f"Error: Permission denied for the file at {path}.")
    except Exception as e:
        raise Exception(f"An unexpected error occurred: {e}")
    
    return buffer



def csv_buffer(path: str) -> Dict[str, str]:
    buffer = {}
    column_number = 2
    try:
        with open(path, 'r', newline='') as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                if len(row) == column_number:
                    buffer[row[0].strip()] = row[1].strip()

    except FileNotFoundError:
        raise FileNotFoundError(f"Error: The file at {path} was not found.")
    except PermissionError:
        raise PermissionError(f"Error: Permission denied for the file at {path}.")
    except csv.Error as e:
        raise csv.Error(f"Error: An error occurred while reading the CSV file at {path}: {e}")
    except Exception as e:
        raise Exception(f"An unexpected error occurred: {e}")

    return buffer


def excel_buffer(path: str) -> Dict[str, str]:
    buffer = {}
    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            #TODO: fix this (buffer is a dict)
            if row[0] is not None and row[1] is not None:
                buffer[row[0].strip()] = str(row[1]).strip()

    except FileNotFoundError:
        raise FileNotFoundError(f"Error: The file at {path} was not found.")
    except PermissionError:
        raise PermissionError(f"Error: Permission denied for the file at {path}. Hint: You must close the file before running the program")
    except InvalidFileException:
        raise InvalidFileException(f"Error: The file at {path} is not a valid Excel file.")
    except Exception as e:
        raise Exception(f"An unexpected error occurred: {e}")

    return buffer


def gsheet_buffer(url) -> List[List[Any]]:
    # TODO: Apply google sheets logic (configure api and get url params)
    raise NotImplementedError("Google Sheets buffer reading is not implemented yet.")
