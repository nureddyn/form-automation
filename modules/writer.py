import os
from openpyxl import Workbook
from .reader import *
from .path_handling import get_absolute_path

def write_to_pdf(buffer: Dict[str, str], template: File):

    reader = PyPDF2.PdfReader(template.path)
    writer = PyPDF2.PdfWriter()

    for page_num in range(len(reader.pages)):
            page = reader.pages[page_num]
            fields = reader.get_fields()

            writer.add_page(page)

            if fields:
                for key in buffer:
                    if key in fields:
                        # fields[key] = {"/V": buffer[key]}
                
                        writer.update_page_form_field_values(writer.pages[page_num], {key: buffer[key]})

        # Write the final pdf form in the output folder
    current_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(current_dir, '..', 'output', "output.pdf")

    with open(output_path, "wb") as output_file:
        writer.write(output_file)


    # with open(template.path, 'rb') as template_file:
    #     reader = PyPDF2.PdfReader(template_file)

    #     writer = PyPDF2.PdfWriter()

    #     for page_num in range(len(reader.pages)):
    #         page = reader.pages[page_num]
    #         fields = reader.get_fields()

    #         writer.add_page(page)

            # if fields:
            #     for key in buffer:
            #         if key in fields:
            #             # fields[key] = {"/V": buffer[key]}
                
            #             writer.update_page_form_field_values(writer.pages[page_num], {key: buffer[key]})

    #     # Write the final pdf form in the output folder
    #     current_dir = os.path.dirname(os.path.abspath(__file__))
    #     output_path = os.path.join(current_dir, '..', 'output', "output.pdf")

    #     with open(output_path, "wb") as output_file:
    #         writer.write(output_file)


def write_to_excel(file: File):
    reader = file_reader(file)
    content = reader.get('fields', {})
    workbook = Workbook()
    sheet = workbook.active

    field_names = list(content.keys())

    for row_index, (key, value) in enumerate(content.items(), start=1):
        sheet.cell(row=row_index, column=1, value=key)
        sheet.cell(row=row_index, column=2, value=value if value is not None else "N/A")

    folder_to_save = get_absolute_path("output")
    file_path = os.path.join(folder_to_save, f"{reader['TYPE']}.xlsx")
    workbook.save(file_path)
