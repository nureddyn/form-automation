import os
from typing import List, Any, Dict
from modules.classes.File import File
import csv
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd
import PyPDF2
import re
from .file_formats import data_formats
"""
Validate file according to process of transport data from a spreadsheet to word/pdf
"""

# TODO: Validate files using a valid_extension (done), a valid_data_file (to check if the content format is correct) and a valid_template_file()

def get_input_file_list() -> List[File]:
    current_directory = os.path.dirname(os.path.realpath(__file__))
    input_directory = os.path.join(current_directory, '..', 'input')
    file_list = os.listdir(input_directory)

    files = []

    for file in file_list:
        full_path = os.path.join(input_directory, file)
    
        file = File(full_path)
        if valid_extension(file) and is_valid_format(file):
            files.append(file)
    return files


def valid_extension(file: File) -> bool:
    if file.extension in [".csv", ".xlsx", ".gsheet", ".pdf"]:
        return True
    # raise ValueError("File extension not allowed")


# TODO: This function must make a distinction between input and template files
def is_valid_format(file: File) -> bool:
    # Read the file and check if the format comply with predefined conventions
    file_buffer = file_reader(file)
    file_buffer_keys = file_buffer.keys()

    for format in data_formats:
        if format.keys() == file_buffer_keys and file_buffer.get('TYPE') == format.get('TYPE'):
            return True
    return False

# TODO: Create a function that determines which template will be used.
def data_file_format(file: File) -> str:
    return

def get_file_type(file: File) -> str:
    if file.extension in [".csv", ".xlsx", ".gsheet"]:
        return 'Data file'
    elif file.extension in [".docx", ".pdf"]:
        return 'Template file'
    raise ValueError("File type not allowed")


# TODO: Pull the most recent forms from source to 'templates' folder
def update_templates():
    return


def generate_forms(input_files: List[File]):

    update_templates()

    # TODO: Ensure the template file can be scanned to get the form type
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(current_dir)

    templates_folder = os.path.join(project_dir, 'templates')
    templates_list = os.listdir(templates_folder)
    template_files = [File(os.path.join(templates_folder, template)) for template in templates_list]
    
    for file in input_files:
        input_buffer = file_reader(file)
        required_form_type = input_buffer.get('TYPE')

        required_template = next((template for template in template_files if file_reader(template).get('TYPE').strip() == required_form_type.strip()), None)
        file_writer(input_buffer, required_template)



def file_reader(file: File) -> Dict[str, str]:
    if file.extension == ".csv":
        return csv_buffer(file.path)
    elif file.extension == ".xlsx":
        return excel_buffer(file.path)
    elif file.extension == ".gsheet":
        return gsheet_buffer(file.path)
    elif file.extension == ".pdf":
        return pdf_buffer(file.path)
    raise ValueError("File type not supported for reading")


# TODO: Implement writer functionality for docx
def file_writer(buffer, template: File):
    try:
        if template.extension == ".pdf":
            write_to_pdf(buffer, template)
        else:
            raise NotImplementedError("writing to docx is not implemented yet.")
    except Exception as e:
        raise Exception(f"{e}")


# 1. Function with test
# TODO: check if implementation is correct
def write_to_pdf(buffer: Dict[str, str], template: File):
    reader = file_reader(template).get('reader')
    writer = PyPDF2.PdfFileWriter()

    for page_num in range(reader.numPages):
        page = reader.getPage(page_num)
        writer.addPage(page)

        fields = reader.getFields(page)
        if fields:
            for key in buffer:
                if key in fields:
                    fields[key].update({"/V": buffer[key]})
            writer.updatePageFormFieldValues(page, fields)

    # Write the final pdf form in input folder
    current_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(current_dir, '..', 'output', "output.pdf")
    
    with open(output_path, "wb") as output_file:
        writer.write(output_file)


# TODO: check if implementation is correct
def pdf_buffer(path: str) -> PyPDF2.PdfReader:
    try:
        with open(path, "rb") as f:
            pdf_reader = PyPDF2.PdfReader(f)
            metadata = pdf_reader.metadata

            first_page = pdf_reader.pages[0].extract_text().splitlines()

            # substring_pattern = re.compile(r'\b\w*FACILIT\w*\b', re.IGNORECASE)
            form_type = [line for line in first_page if re.search(r'\b\w*FACILIT\w*\b', line, re.IGNORECASE)][0]

            buffer = {'METADATA': metadata, 'reader': pdf_reader, 'TYPE': form_type}

    except FileNotFoundError:
        raise FileNotFoundError(f"Error: The file at {path} was not found.")
    except PermissionError:
        raise PermissionError(f"Error: Permission denied for the file at {path}.")
    except Exception as e:
        raise Exception(f"An unexpected error occurred: {e}")
    
    return buffer



def csv_buffer(path: str) -> Dict[str, str]:
    buffer = {}
    column_number = 2
    try:
        with open(path, 'r', newline='') as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                if len(row) == column_number:
                    buffer[row[0].strip()] = row[1].strip()

    except FileNotFoundError:
        raise FileNotFoundError(f"Error: The file at {path} was not found.")
    except PermissionError:
        raise PermissionError(f"Error: Permission denied for the file at {path}.")
    except csv.Error as e:
        raise csv.Error(f"Error: An error occurred while reading the CSV file at {path}: {e}")
    except Exception as e:
        raise Exception(f"An unexpected error occurred: {e}")

    return buffer


def excel_buffer(path: str) -> Dict[str, str]:
    buffer = {}
    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            #TODO: fix this (buffer is a dict)
            if row[0] is not None and row[1] is not None:
                buffer[row[0].strip()] = str(row[1]).strip()

    except FileNotFoundError:
        raise FileNotFoundError(f"Error: The file at {path} was not found.")
    except PermissionError:
        raise PermissionError(f"Error: Permission denied for the file at {path}. Hint: You must close the file before running the program")
    except InvalidFileException:
        raise InvalidFileException(f"Error: The file at {path} is not a valid Excel file.")
    except Exception as e:
        raise Exception(f"An unexpected error occurred: {e}")

    return buffer    

def gsheet_buffer(url) -> List[List[Any]]:
    # TODO: Apply google sheets logic (configure api and get url params)
    raise NotImplementedError("Google Sheets buffer reading is not implemented yet.")
